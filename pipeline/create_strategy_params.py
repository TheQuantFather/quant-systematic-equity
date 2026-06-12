"""
create_strategy_params.py — Create/reset the strategy_params.xlsx template.

The Python file is the source of truth; re-running overwrites the workbook.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


from config import PARAMS_FILE as OUT, MODELS_REF
from utils import get_logger

log = get_logger("create_strategy_params")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
BORDER_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                     top=BORDER_SIDE,  bottom=BORDER_SIDE)

# Category label fills for grouping strategies in the Constraints sheet
CAT_FILLS = {
    "core_active":         PatternFill("solid", fgColor="E3F2FD"),
    "core_active_strict":  PatternFill("solid", fgColor="BBDEFB"),
    "abs_return":          PatternFill("solid", fgColor="E8F5E9"),
}


def _header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _data_row(ws, row, cols, sid=None):
    fill = CAT_FILLS.get(sid) if sid else None
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Strategies ───────────────────────────────────────────────────────

def build_strategies(wb):
    ws = wb.create_sheet("Strategies")
    headers = [
        "strategy_id", "name", "active",
        "benchmark_file", "benchmark_index", "universe_index",
        "solver", "objective", "risk_aversion", "investable_universe", "description",
    ]
    # risk_aversion: objective parameter for maximize_alpha. 0 = pure linear alpha
    # (loads conviction tilts). >0 adds a mean-variance penalty on active variance
    # (w-b)'Σ(w-b), shrinking bets toward the benchmark. Left at 0 — walk-forward
    # backtests show a positive penalty hurts the current ALP001 alpha.
    data = [
        # id                   name                    active  benchmark_file benchmark_index universe_index solver  objective          risk_aversion universe          description
        ["core_active",        "Core Active",          "TRUE", "",            "sp500",        "",            "MOSEK", "maximize_alpha",  "0",          "benchmark_only", "Live-sized core active portfolio: benchmark investable universe, 55-60 names, whole-share sizing for a 35k USD account, and 5% TE."],
        ["core_active_strict", "Core Active (Strict)", "FALSE","",            "sp500",        "",            "MOSEK", "maximize_alpha",  "0",          "benchmark_only", "Core active constraint set with stricter 2% TE, ±1% active exposure bands, and a 4% issuer cap."],
        ["abs_return",        "Absolute Return",      "TRUE", "",            "sp500",        "sp500",       "MOSEK", "maximize_alpha",  "0",          "universe",       "Maximize alpha with core-like exposure bands: 35-45 names, benchmark-relative sector/industry, cap-bucket diversification, and 12% active-risk cap."],
    ]

    ws.append(headers)
    _header(ws, 1, len(headers))
    for i, row in enumerate(data, start=2):
        ws.append(row)
        _data_row(ws, i, len(headers), sid=row[0])

    _widths(ws, [20, 22, 7, 28, 18, 18, 10, 18, 13, 18, 60])
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── Sheet 2: Constraints ──────────────────────────────────────────────────────

CORE_CONSTRAINTS = [
    ["long_only",                  "TRUE",  "TRUE",  "No short positions"],
    ["fully_invested",             "TRUE",  "TRUE",  "Weights sum to 1"],
    ["max_active_risk",            "0.05",  "TRUE",  "Max annual tracking error 5%"],
    ["max_stock_active_weight",    "0.05",  "TRUE",  "Max ±5% active weight per stock; TE remains binding"],
    ["max_sector_active_weight",   "0.03",  "FALSE", "Disabled while testing whole-share live portfolio feasibility"],
    ["max_industry_active_weight", "0.03",  "FALSE", "Disabled while testing whole-share live portfolio feasibility"],
    ["min_positions",             "55",    "TRUE",  "Minimum 55 securities for lower tracking error with whole-share sizing"],
    ["max_positions",             "60",    "TRUE",  "Maximum 60 securities for 30k live-sized whole-share portfolio"],
    ["min_position_if_held",       "0.01",  "TRUE",  "Minimum 1% if selected; 400 EUR trading minimum handled at broker sizing"],
    ["portfolio_value_usd",        "35000", "TRUE",  "Approximate live/paper account value for whole-share lot-size optimisation"],
    ["max_cash_weight",            "0.005", "TRUE",  "Max 0.5% residual cash in whole-share optimisation"],
    ["lot_size_max_overweight",    "0.03",  "TRUE",  "Allow up to 3% extra per-name weight from whole-share lot rounding"],
    ["use_lp_prescreen",           "FALSE", "FALSE", "Disabled for live lot-size optimisation; full universe is cleaner while exploratory"],
    ["lp_prescreen_multiplier",    "10",    "FALSE", "Unused while pre-screen is disabled"],
    ["excluded_tickers",           "AL",    "TRUE",  "Exclude untradable/delisted tickers from live-sized optimisation"],
    ["large_cap_min_market_cap",   "10000000000", "TRUE", "Large cap threshold: >= $10B"],
    ["mid_cap_min_market_cap",     "2000000000",  "TRUE", "Mid cap threshold: >= $2B and < $10B"],
    ["max_large_cap_position",     "0.05",  "TRUE",  "Max 5% per large-cap stock"],
    ["min_large_cap_weight",       "0.80",  "TRUE",  "Benchmark-only live sizing: min 80% large-cap exposure"],
    ["max_large_cap_weight",       "1.00",  "TRUE",  "Benchmark-only live sizing: allow up to 100% large-cap exposure"],
    ["max_mid_cap_position",       "0.04",  "TRUE",  "Max 4% per mid-cap stock"],
    ["min_mid_cap_weight",         "0.00",  "FALSE", "No mid-cap floor for benchmark-only live sizing"],
    ["max_mid_cap_weight",         "0.15",  "TRUE",  "Benchmark-only live sizing: max 15% mid-cap exposure"],
    ["max_small_cap_position",     "0.03",  "TRUE",  "Max 3% per small-cap or missing-cap stock"],
    ["min_small_cap_weight",       "0.00",  "FALSE", "No small-cap floor for benchmark-only live sizing"],
    ["max_small_cap_weight",       "0.10",  "TRUE",  "Benchmark-only live sizing: max 10% small-cap exposure"],
]

STRICT_OVERRIDES = {
    "max_active_risk":            ["0.02", "TRUE", "Max annual tracking error 2%"],
    "max_stock_active_weight":    ["0.01", "TRUE", "Max ±1% active weight per stock"],
    "max_sector_active_weight":   ["0.01", "TRUE", "Max ±1% active weight per GICS sector"],
    "max_industry_active_weight": ["0.01", "TRUE", "Max ±1% active weight per industry"],
}

STRICT_EXTRA_CONSTRAINTS = [
    ["max_issuer_weight", "0.04", "TRUE", "Max 4% total weight per issuer across share classes"],
]

ABS_RETURN_CONSTRAINTS = [
    ["long_only",                  "TRUE",  "TRUE",  "No short positions"],
    ["fully_invested",             "FALSE", "TRUE",  "Cash allowed up to max_cash_weight; held only when the vol cap binds"],
    ["max_cash_weight",            "0.20",  "TRUE",  "Max 20% cash; de-risk headroom for max_portfolio_vol (sector bands cap cash at ~22% anyway)"],
    ["max_active_risk",            "0.12",  "TRUE",  "Max 12% annual active risk vs configured benchmark"],
    ["max_position",               "0.04",  "TRUE",  "Global max 4% per stock; bucket caps tighten mid/small further"],
    ["min_positions",              "35",    "TRUE",  "Minimum 35 securities for retail-scale diversification"],
    ["max_positions",              "45",    "TRUE",  "Maximum 45 securities initially"],
    ["min_position_if_held",        "0.01",  "TRUE",  "Minimum 1% if selected; avoids small live-account positions"],
    ["max_sector_active_weight",    "0.02",  "TRUE",  "Sector weight must stay within ±2% of configured benchmark"],
    ["max_industry_active_weight",  "0.02",  "TRUE",  "Industry weight must stay within ±2% of configured benchmark"],
    ["equal_sector_weight",         "FALSE", "FALSE", "Disabled: conflicts with benchmark-relative sector exposure"],
    ["sector_weight_tolerance",     "0.01",  "FALSE", "Unused while equal_sector_weight is disabled"],
    ["max_portfolio_vol",           "0.17",  "TRUE",  "Max 17% annual vol; binds in stressed regimes only (ex-ante median ~16.3%)"],
    ["max_industry_weight",         "0.10",  "FALSE", "Disabled: industry exposure is controlled vs benchmark instead"],
    ["use_alpha_prescreen",         "FALSE", "FALSE", "Do not pre-screen the Sharpe MIP; solve on the real universe"],
    ["large_cap_min_market_cap",    "10000000000", "TRUE", "Large cap threshold: >= $10B"],
    ["mid_cap_min_market_cap",      "2000000000",  "TRUE", "Mid cap threshold: >= $2B and < $10B"],
    ["max_large_cap_position",      "0.04",  "TRUE",  "Max 4% per large-cap stock"],
    ["min_large_cap_weight",        "0.60",  "TRUE",  "Large-cap sleeve floor: 60%"],
    ["max_large_cap_weight",        "0.75",  "TRUE",  "Large-cap sleeve cap: 75%"],
    ["max_mid_cap_position",        "0.03",  "TRUE",  "Max 3% per mid-cap stock"],
    ["min_mid_cap_weight",          "0.15",  "TRUE",  "Mid-cap sleeve floor: 15%"],
    ["max_mid_cap_weight",          "0.30",  "TRUE",  "Mid-cap sleeve cap: 30%"],
    ["max_small_cap_position",      "0.02",  "TRUE",  "Max 2% per small-cap or missing-cap stock"],
    ["min_small_cap_weight",        "0.00",  "TRUE",  "Small-cap sleeve floor: 0%"],
    ["max_small_cap_weight",        "0.10",  "TRUE",  "Small-cap sleeve cap: 10%"],
]


def _constraint_rows(strategy_id: str, constraints: list[list]) -> list[list]:
    return [[strategy_id, *row] for row in constraints]


def _strict_constraints() -> list[list]:
    rows = []
    for name, value, enabled, notes in CORE_CONSTRAINTS:
        rows.append([name, *STRICT_OVERRIDES.get(name, [value, enabled, notes])])
        if name == "max_stock_active_weight":
            rows.extend(STRICT_EXTRA_CONSTRAINTS)
    return rows


def build_constraints(wb):
    ws = wb.create_sheet("Constraints")
    headers = ["strategy_id", "constraint", "value", "enabled", "notes"]
    data = (
        _constraint_rows("core_active", CORE_CONSTRAINTS)
        + _constraint_rows("core_active_strict", _strict_constraints())
        + _constraint_rows("abs_return", ABS_RETURN_CONSTRAINTS)
    )

    ws.append(headers)
    _header(ws, 1, len(headers))
    for i, row in enumerate(data, start=2):
        ws.append(row)
        _data_row(ws, i, len(headers), sid=row[0])

    _widths(ws, [22, 30, 12, 9, 65])
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── Sheet 3: Alpha_Weights ────────────────────────────────────────────────────

def build_alpha_weights(wb):
    ws = wb.create_sheet("Alpha_Weights")
    headers = ["strategy_id", "model_id", "model_name", "weight", "notes"]
    data = [
        # core_active — composite
        ["core_active",        "ALP001",  "Alpha Composite", "1.0", "Equal-weight composite of all base models"],
        ["core_active_strict", "ALP001",  "Alpha Composite", "1.0", "Same composite alpha, tighter risk budget"],
        # abs_return — composite
        ["abs_return",        "ALP001",  "Alpha Composite", "1.0", "Composite alpha for Sharpe optimisation"],
    ]

    ws.append(headers)
    _header(ws, 1, len(headers))
    for i, row in enumerate(data, start=2):
        ws.append(row)
        _data_row(ws, i, len(headers), sid=row[0])

    _widths(ws, [22, 12, 20, 9, 60])
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── Sheet 4: Reference ────────────────────────────────────────────────────────

def _model_ref_rows() -> list[list]:
    """Build model reference table rows from models_reference.csv."""
    df = pd.read_csv(MODELS_REF)[["ModelID", "Model", "IsComposite"]].drop_duplicates()
    rows: list[list] = [["Model ID", "Name", "Type"]]
    for _, r in df.iterrows():
        rows.append([r["ModelID"], r["Model"], "Composite" if int(r["IsComposite"]) else "Base"])
    return rows


def build_reference(wb):
    ws = wb.create_sheet("Reference")

    def section(row: int, title: str) -> None:
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, size=12)

    def table(start_row: int, rows: list[list]) -> int:
        for r_idx, row_data in enumerate(rows, start=start_row):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == start_row:
                    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                if r_idx > start_row and r_idx % 2 == 0:
                    cell.fill = ALT_FILL
        return start_row + len(rows)  # next available row

    r = 1
    section(r, "Strategies overview")
    r = table(r + 2, [
        ["strategy_id",        "Objective",           "Alpha signal",                          "Universe",      "Investor profile"],
        ["core_active",        "maximize_alpha",       "Composite (all factors)",               "Benchmark",     "Institutional — benchmark-aware"],
        ["core_active_strict", "maximize_alpha",       "Composite (all factors)",               "Benchmark",     "Strict benchmark-aware core with issuer cap"],
        ["abs_return",         "maximize_alpha",       "Composite (all factors)",               "Full universe", "Absolute return — alpha with benchmark-aware exposures"],
    ])

    r += 2
    section(r, "Models available for Alpha_Weights")
    r = table(r + 2, _model_ref_rows())

    r += 2
    section(r, "Objectives")
    r = table(r + 2, [
        ["Objective",         "Description"],
        ["maximize_alpha",    "Benchmark-aware: maximize active alpha (tracking-error constrained). Needs benchmark_index or benchmark_file."],
    ])

    r += 2
    section(r, "Constraint reference")
    table(r + 2, [
        ["Constraint",               "Applies to",               "Description"],
        ["max_active_risk",          "maximize_alpha",            "Max annual tracking error vs benchmark"],
        ["max_stock_active_weight",  "maximize_alpha",            "Max ±active weight per stock"],
        ["max_sector_active_weight", "all objectives with benchmark", "Max ±active weight per GICS sector"],
        ["max_industry_active_weight","all objectives with benchmark","Max ±active weight per industry"],
        ["min_positions",            "integer objectives",        "Minimum number of held securities"],
        ["max_positions",            "integer objectives",        "Maximum number of held securities"],
        ["min_position_if_held",      "integer objectives",        "Minimum weight when a security is selected"],
        ["max_large_cap_position",   "all objectives",            "Max absolute weight per large-cap stock"],
        ["min_large_cap_weight",     "all objectives",            "Min total weight in large-cap bucket"],
        ["max_large_cap_weight",     "all objectives",            "Max total weight in large-cap bucket"],
        ["max_mid_cap_position",     "all objectives",            "Max absolute weight per mid-cap stock"],
        ["min_mid_cap_weight",       "all objectives",            "Min total weight in mid-cap bucket"],
        ["max_mid_cap_weight",       "all objectives",            "Max total weight in mid-cap bucket"],
        ["max_small_cap_position",   "all objectives",            "Max absolute weight per small-cap or missing-cap stock"],
        ["min_small_cap_weight",     "all objectives",            "Min total weight in small-cap bucket"],
        ["max_small_cap_weight",     "all objectives",            "Max total weight in small-cap bucket"],
        ["max_position",             "all objectives",            "Max absolute weight per stock"],
        ["max_sector_weight",        "all objectives",            "Max absolute weight per GICS sector"],
        ["min_sector_weight",        "all objectives",            "Min absolute weight per non-excluded sector"],
        ["equal_sector_weight",      "all objectives",            "Each sector gets 1/n_sectors ± sector_weight_tolerance"],
        ["sector_weight_tolerance",  "all objectives",            "±tolerance around equal-weight target (e.g. 0.01 = ±1%)"],
        ["excluded_sectors",         "all objectives",            "Pipe-separated sectors to zero out, e.g. Energy|Materials"],
        ["max_issuer_weight",        "all objectives",            "Max absolute weight per issuer across share classes (CIK-based when available)"],
        ["max_industry_weight",      "all objectives",            "Max absolute weight per SimFin industry group"],
        ["max_portfolio_vol",        "maximize_alpha",            "Max annual portfolio volatility"],
        ["max_cash_weight",          "maximize_alpha",            "Max cash weight; weights sum in [1 - cash, 1]. Disable to force fully invested"],
    ])

    _widths(ws, [28, 28, 75])


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_strategies(wb)
    build_constraints(wb)
    build_alpha_weights(wb)
    build_reference(wb)

    wb.save(OUT)
    log.info("Created %s  (%s KB)", OUT, OUT.stat().st_size // 1024)


if __name__ == "__main__":
    main()
