"""
create_strategy_params.py — Create/reset the strategy_params.xlsx template.

Run once to create, then edit in Excel.
Re-running will OVERWRITE any manual edits — only use to reset.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import PARAMS_FILE as OUT
from utils import get_logger

log = get_logger("create_strategy_params")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
BORDER_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                     top=BORDER_SIDE,  bottom=BORDER_SIDE)

# Category label fills for grouping strategies in the Constraints sheet
CAT_FILLS = {
    "core_active":         PatternFill("solid", fgColor="E3F2FD"),
    "core_active_strict":  PatternFill("solid", fgColor="BBDEFB"),
    "abs_return":          PatternFill("solid", fgColor="E8F5E9"),
    "min_variance":        PatternFill("solid", fgColor="FFF3E0"),
    "quality_compounder":  PatternFill("solid", fgColor="F3E5F5"),
    "defensive":           PatternFill("solid", fgColor="FCE4EC"),
    "value_hunt":          PatternFill("solid", fgColor="E0F7FA"),
    "momentum":            PatternFill("solid", fgColor="FFF8E1"),
    "all_weather":         PatternFill("solid", fgColor="E8EAF6"),
}


def _header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _data_row(ws, row, cols, sid=None):
    fill = CAT_FILLS.get(sid) if sid else None
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Strategies ───────────────────────────────────────────────────────

def build_strategies(wb):
    ws = wb.create_sheet("Strategies")
    headers = [
        "strategy_id", "name", "active",
        "benchmark_file", "alpha_model", "alpha_date", "risk_date",
        "solver", "objective", "investable_universe", "description",
    ]
    data = [
        # id               name                       active  benchmark                  alpha    a_date       r_date       solver     objective           universe         description
        ["core_active",    "Core Active",             "TRUE", "MSCI_USA_2026_05_07.csv", "ALP001","2026-04-01","2026-04-01","CLARABEL","maximize_alpha",   "benchmark_only","Benchmark-aware. Max 4% active risk vs MSCI USA. Composite alpha."],
        ["core_active_strict","Core Active (Strict)", "TRUE", "MSCI_USA_2026_05_07.csv", "ALP001","2026-04-01","2026-04-01","CLARABEL","maximize_alpha",   "benchmark_only","Tighter benchmark-aware. Max 2% active risk, ±1% per stock and sector."],
        ["abs_return",     "Absolute Return",         "TRUE", "MSCI_USA_2026_05_07.csv", "ALP001","2026-04-01","2026-04-01","CLARABEL","maximize_sharpe",  "universe",      "Maximize Sharpe. Equal sector weight ±1%, max 5% per stock, 18% vol cap."],
        ["min_variance",   "Minimum Variance",        "TRUE", "MSCI_USA_2026_05_07.csv", "ALP001","2026-04-01","2026-04-01","CLARABEL","minimize_variance","universe",      "Pure risk minimisation — no alpha signal. Capital preservation mandate."],
        ["quality_compounder","Quality Compounder",   "TRUE", "MSCI_USA_2026_05_07.csv", "QUAL001","2026-04-01","2026-04-01","CLARABEL","maximize_sharpe", "universe",      "Quality-only alpha. Buffett-style durable businesses. Excludes Energy & Materials."],
        ["defensive",      "Defensive Income",        "TRUE", "MSCI_USA_2026_05_07.csv", "QUAL001","2026-04-01","2026-04-01","CLARABEL","maximize_sharpe",  "universe",      "Quality + Low-Vol blend. Low vol cap (12%), sector floors to stay diversified."],
        ["value_hunt",     "Value Hunt",              "TRUE", "MSCI_USA_2026_05_07.csv", "VAL001","2026-04-01","2026-04-01","CLARABEL","maximize_alpha",   "benchmark_only","Value-only alpha vs MSCI USA benchmark. Wider active risk budget (6%)."],
        ["momentum",       "Momentum",                "TRUE", "MSCI_USA_2026_05_07.csv", "MOM001","2026-04-01","2026-04-01","CLARABEL","maximize_sharpe",  "universe",      "Momentum-only alpha. Chases what's working. Higher vol tolerance (20%)."],
        ["all_weather",    "All-Weather GARP",        "TRUE", "MSCI_USA_2026_05_07.csv", "ALP001","2026-04-01","2026-04-01","CLARABEL","maximize_sharpe",  "universe",      "Quality + Growth + Value blend. Equal sector weight ±2%. Balanced core holding."],
    ]

    ws.append(headers)
    _header(ws, 1, len(headers))
    for i, row in enumerate(data, start=2):
        ws.append(row)
        _data_row(ws, i, len(headers), sid=row[0])

    _widths(ws, [20, 22, 7, 28, 12, 12, 12, 10, 18, 18, 60])
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── Sheet 2: Constraints ──────────────────────────────────────────────────────

def build_constraints(wb):
    ws = wb.create_sheet("Constraints")
    headers = ["strategy_id", "constraint", "value", "enabled", "notes"]
    data = [
        # ── core_active ──────────────────────────────────────────────────────
        ["core_active", "long_only",                   "TRUE", "TRUE",  "No short positions"],
        ["core_active", "fully_invested",              "TRUE", "TRUE",  "Weights sum to 1"],
        ["core_active", "max_active_risk",             "0.04", "TRUE",  "Max annual tracking error (4%)"],
        ["core_active", "max_stock_active_weight",     "0.02", "TRUE",  "Max ±2% active weight per stock"],
        ["core_active", "max_sector_active_weight",    "0.02", "TRUE",  "Max ±2% active weight per GICS sector"],
        ["core_active", "max_industry_active_weight",  "0.02", "TRUE",  "Max ±2% active weight per industry"],

        # ── core_active_strict ───────────────────────────────────────────────
        ["core_active_strict", "long_only",                  "TRUE", "TRUE",  "No short positions"],
        ["core_active_strict", "fully_invested",             "TRUE", "TRUE",  "Weights sum to 1"],
        ["core_active_strict", "max_active_risk",            "0.02", "TRUE",  "Max 2% annual tracking error"],
        ["core_active_strict", "max_stock_active_weight",    "0.01", "TRUE",  "Max ±1% active weight per stock"],
        ["core_active_strict", "max_sector_active_weight",   "0.01", "TRUE",  "Max ±1% active weight per GICS sector"],
        ["core_active_strict", "max_industry_active_weight", "0.01", "TRUE",  "Max ±1% active weight per industry"],

        # ── abs_return ───────────────────────────────────────────────────────
        ["abs_return",  "long_only",                   "TRUE", "TRUE",  "No short positions"],
        ["abs_return",  "fully_invested",              "TRUE", "TRUE",  "Weights sum to 1"],
        ["abs_return",  "max_position",                "0.05", "TRUE",  "Max 5% per stock"],
        ["abs_return",  "equal_sector_weight",         "TRUE", "TRUE",  "Each sector gets 1/11 of portfolio"],
        ["abs_return",  "sector_weight_tolerance",     "0.01", "TRUE",  "±1% around equal sector target"],
        ["abs_return",  "max_portfolio_vol",           "0.18", "TRUE",  "Max 18% annual portfolio volatility"],
        ["abs_return",  "max_industry_weight",         "0.10", "TRUE",  "Max 10% in any single industry"],

        # ── min_variance ─────────────────────────────────────────────────────
        ["min_variance","long_only",                   "TRUE", "TRUE",  "No short positions"],
        ["min_variance","fully_invested",              "TRUE", "TRUE",  "Weights sum to 1"],
        ["min_variance","max_position",                "0.05", "TRUE",  "Max 5% per stock — prevents degenerate concentration"],
        ["min_variance","max_sector_weight",           "0.25", "TRUE",  "Max 25% in any single sector"],
        ["min_variance","max_industry_weight",         "0.10", "TRUE",  "Max 10% in any single industry"],

        # ── quality_compounder ───────────────────────────────────────────────
        ["quality_compounder","long_only",             "TRUE", "TRUE",  "No short positions"],
        ["quality_compounder","fully_invested",        "TRUE", "TRUE",  "Weights sum to 1"],
        ["quality_compounder","max_position",          "0.08", "TRUE",  "Max 8% per stock — more concentrated than abs_return"],
        ["quality_compounder","max_sector_weight",     "0.25", "TRUE",  "Max 25% in any single sector"],
        ["quality_compounder","excluded_sectors",      "Energy|Materials","TRUE","Exclude cyclical capital-intensive sectors"],
        ["quality_compounder","max_portfolio_vol",     "0.16", "TRUE",  "Max 16% annual volatility"],

        # ── defensive ────────────────────────────────────────────────────────
        ["defensive",   "long_only",                   "TRUE", "TRUE",  "No short positions"],
        ["defensive",   "fully_invested",              "TRUE", "TRUE",  "Weights sum to 1"],
        ["defensive",   "max_position",                "0.04", "TRUE",  "Max 4% per stock"],
        ["defensive",   "min_sector_weight",           "0.05", "TRUE",  "Min 5% in every sector — prevents sector exclusion"],
        ["defensive",   "max_sector_weight",           "0.20", "TRUE",  "Max 20% in any sector"],
        ["defensive",   "max_portfolio_vol",           "0.12", "TRUE",  "Max 12% annual vol — low-risk mandate"],
        ["defensive",   "max_industry_weight",         "0.08", "TRUE",  "Max 8% per industry"],

        # ── value_hunt ───────────────────────────────────────────────────────
        ["value_hunt",  "long_only",                   "TRUE", "TRUE",  "No short positions"],
        ["value_hunt",  "fully_invested",              "TRUE", "TRUE",  "Weights sum to 1"],
        ["value_hunt",  "max_active_risk",             "0.06", "TRUE",  "Max 6% active risk — wider than core_active"],
        ["value_hunt",  "max_stock_active_weight",     "0.03", "TRUE",  "Max ±3% active weight per stock"],
        ["value_hunt",  "max_sector_active_weight",    "0.04", "TRUE",  "Max ±4% sector active weight — allow value sector tilts"],

        # ── momentum ─────────────────────────────────────────────────────────
        ["momentum",    "long_only",                   "TRUE", "TRUE",  "No short positions"],
        ["momentum",    "fully_invested",              "TRUE", "TRUE",  "Weights sum to 1"],
        ["momentum",    "max_position",                "0.05", "TRUE",  "Max 5% per stock"],
        ["momentum",    "max_sector_weight",           "0.30", "TRUE",  "Max 30% per sector — momentum clusters in sectors"],
        ["momentum",    "max_portfolio_vol",           "0.20", "TRUE",  "Max 20% vol — momentum accepts higher risk"],

        # ── all_weather ──────────────────────────────────────────────────────
        ["all_weather", "long_only",                   "TRUE", "TRUE",  "No short positions"],
        ["all_weather", "fully_invested",              "TRUE", "TRUE",  "Weights sum to 1"],
        ["all_weather", "max_position",                "0.04", "TRUE",  "Max 4% per stock"],
        ["all_weather", "equal_sector_weight",         "TRUE", "TRUE",  "Equal weight across all 11 GICS sectors"],
        ["all_weather", "sector_weight_tolerance",     "0.02", "TRUE",  "±2% around equal sector target"],
        ["all_weather", "max_portfolio_vol",           "0.15", "TRUE",  "Max 15% annual volatility"],
        ["all_weather", "max_industry_weight",         "0.08", "TRUE",  "Max 8% per industry"],
    ]

    ws.append(headers)
    _header(ws, 1, len(headers))
    for i, row in enumerate(data, start=2):
        ws.append(row)
        _data_row(ws, i, len(headers), sid=row[0])

    _widths(ws, [22, 30, 12, 9, 65])
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── Sheet 3: Alpha_Weights ────────────────────────────────────────────────────

def build_alpha_weights(wb):
    ws = wb.create_sheet("Alpha_Weights")
    headers = ["strategy_id", "model_id", "model_name", "weight", "notes"]
    data = [
        # core_active — composite
        ["core_active",        "ALP001",  "Alpha Composite", "1.0", "Equal-weight composite of all base models"],
        ["core_active_strict", "ALP001",  "Alpha Composite", "1.0", "Same composite alpha, tighter risk budget"],
        # abs_return — composite
        ["abs_return",        "ALP001",  "Alpha Composite", "1.0", "Composite alpha for Sharpe optimisation"],
        # min_variance — alpha unused but required as placeholder
        ["min_variance",      "ALP001",  "Alpha Composite", "1.0", "Placeholder — alpha is ignored for minimize_variance"],
        # quality_compounder — pure quality
        ["quality_compounder","QUAL001", "Quality",         "1.0", "Profitability, cash flow quality, leverage"],
        # defensive — quality + low vol blend
        ["defensive",         "QUAL001", "Quality",         "0.5", "Half-weight quality signal"],
        ["defensive",         "LVOL001", "Low Volatility",  "0.5", "Half-weight low-vol signal"],
        # value_hunt — pure value
        ["value_hunt",        "VAL001",  "Value",           "1.0", "Earnings yield, P/B, P/S, P/CF, EV/EBIT"],
        # momentum — pure momentum
        ["momentum",          "MOM001",  "Momentum",        "1.0", "12m and 6m price momentum"],
        # all_weather — quality + growth + value equal blend
        ["all_weather",       "QUAL001", "Quality",         "1.0", "One-third: profitability and balance sheet strength"],
        ["all_weather",       "GRO001",  "Growth",          "1.0", "One-third: earnings, revenue, asset growth"],
        ["all_weather",       "VAL001",  "Value",           "1.0", "One-third: valuation multiples"],
    ]

    ws.append(headers)
    _header(ws, 1, len(headers))
    for i, row in enumerate(data, start=2):
        ws.append(row)
        _data_row(ws, i, len(headers), sid=row[0])

    _widths(ws, [22, 12, 20, 9, 60])
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── Sheet 4: Reference ────────────────────────────────────────────────────────

def build_reference(wb):
    ws = wb.create_sheet("Reference")

    def section(title_cell, title):
        ws[title_cell] = title
        ws[title_cell].font = Font(bold=True, size=12)

    def table(start_row, rows):
        for r_idx, row in enumerate(rows, start=start_row):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == start_row:
                    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                if r_idx > start_row and r_idx % 2 == 0:
                    cell.fill = ALT_FILL

    section("A1", "Strategies overview")
    table(3, [
        ["strategy_id",        "Objective",           "Alpha signal",               "Universe",      "Investor profile"],
        ["core_active",        "maximize_alpha",       "Composite (all factors)",    "Benchmark",     "Institutional — benchmark-aware"],
        ["abs_return",         "maximize_sharpe",      "Composite (all factors)",    "Full universe", "Absolute return — equal sector weight"],
        ["min_variance",       "minimize_variance",    "None",                       "Full universe", "Capital preservation / retirees"],
        ["quality_compounder", "maximize_sharpe",      "Quality only",               "Full universe", "Long-term buy-and-hold, Buffett-style"],
        ["defensive",          "maximize_sharpe",      "Quality + Low Vol (50/50)",  "Full universe", "Conservative income — low drawdown"],
        ["value_hunt",         "maximize_alpha",       "Value only",                 "Benchmark",     "Contrarian / deep-value investors"],
        ["momentum",           "maximize_sharpe",      "Momentum only",              "Full universe", "Growth / trend-following investors"],
        ["all_weather",        "maximize_sharpe",      "Quality + Growth + Value",   "Full universe", "Balanced core — GARP approach"],
    ])

    section("A14", "Models available for Alpha_Weights")
    table(16, [
        ["Model ID", "Name",            "Type",      "Description"],
        ["ALP001",   "Alpha Composite", "Composite", "Equal-weight of all 5 base models"],
        ["QUAL001",  "Quality",         "Base",      "Profitability, cash flow, leverage"],
        ["VAL001",   "Value",           "Base",      "Earnings yield, P/B, P/S, P/CF, EV/EBIT"],
        ["GRO001",   "Growth",          "Base",      "Asset, earnings, revenue, equity, CF growth"],
        ["MOM001",   "Momentum",        "Base",      "12m and 6m price momentum"],
        ["SIZ001",   "Size",            "Base",      "Log market cap (larger = higher score)"],
        ["LVOL001",  "Low Volatility",  "Base",      "Realized vol (lower = higher score)"],
        ["LIQ001",   "Liquidity",       "Base",      "Amihud illiquidity (lower = more liquid = higher score)"],
    ])

    section("A27", "Objectives")
    table(29, [
        ["Objective",         "Description"],
        ["maximize_alpha",    "Benchmark-aware: maximize active alpha (tracking-error constrained). Needs benchmark_file."],
        ["maximize_sharpe",   "Absolute return: maximize Sharpe via Charnes-Cooper transform. No benchmark needed."],
        ["minimize_variance", "Pure risk minimisation: ignores alpha, finds lowest-vol portfolio. No benchmark needed."],
    ])

    section("A35", "Constraint reference")
    table(37, [
        ["Constraint",               "Applies to",               "Description"],
        ["max_active_risk",          "maximize_alpha",            "Max annual tracking error vs benchmark"],
        ["max_stock_active_weight",  "maximize_alpha",            "Max ±active weight per stock"],
        ["max_sector_active_weight", "maximize_alpha",            "Max ±active weight per GICS sector"],
        ["max_industry_active_weight","maximize_alpha",           "Max ±active weight per industry"],
        ["max_position",             "maximize_sharpe, min_var",  "Max absolute weight per stock"],
        ["max_sector_weight",        "all objectives",            "Max absolute weight per GICS sector"],
        ["min_sector_weight",        "all objectives",            "Min absolute weight per non-excluded sector"],
        ["equal_sector_weight",      "all objectives",            "Each sector gets 1/n_sectors ± sector_weight_tolerance"],
        ["sector_weight_tolerance",  "all objectives",            "±tolerance around equal-weight target (e.g. 0.01 = ±1%)"],
        ["excluded_sectors",         "all objectives",            "Pipe-separated sectors to zero out, e.g. Energy|Materials"],
        ["max_industry_weight",      "all objectives",            "Max absolute weight per SimFin industry group"],
        ["max_portfolio_vol",        "maximize_sharpe",           "Max annual portfolio volatility"],
    ])

    _widths(ws, [28, 28, 75])


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_strategies(wb)
    build_constraints(wb)
    build_alpha_weights(wb)
    build_reference(wb)

    wb.save(OUT)
    log.info("Created %s  (%s KB)", OUT, OUT.stat().st_size // 1024)


if __name__ == "__main__":
    main()
